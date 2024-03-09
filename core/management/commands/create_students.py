"""
Create an initial data for students
"""
import random
# import string
# from faker import Faker
from datetime import datetime
from typing import Optional, Any
from django.core.management.base import BaseCommand, CommandParser
from django.db.utils import IntegrityError
import openpyxl

from core.models import (
    Student, Class, AcademicYear,
    AcademicTerm, StudentClass,
    StudentFeeGroup,
    Fee
    # User
)
from utils.utils import generate_random_string, get_initials


class Command(BaseCommand):
    help = "Create Students from initial data"

    def add_arguments(self, parser: CommandParser) -> None:
        return super().add_arguments(parser)

    def handle(self, *args: Any, **options: Any) -> Optional[str]:
        """Create students and Classes/levels, then add students"""

        student_dict, fee_group = extract_excel_data()
        for class_name, students in student_dict.items():
            for row in students:
                acad_year, _ = AcademicYear.objects.get_or_create(
                    year=row["academic_year"]
                )
                acad_term, _ = AcademicTerm.objects.get_or_create(
                    academic_year=acad_year,
                    term="First Term"
                )
                my_class, _ = Class.objects.get_or_create(
                    name=class_name
                )
                middle_name = ""
                if row["middle_name"] is None:
                    middle_name = ""
                studentID = generate_random_string(length=8) + get_initials(
                        row["first_name"], middle_name, row["last_name"]
                )

                try:
                    student = Student.objects.get(
                        first_name=row["first_name"],
                        last_name=row["last_name"],
                        middle_name=middle_name
                    )
                except Student.DoesNotExist:
                    try:
                        student = Student.objects.create(
                            student_id=studentID,
                            gender=random.choice(["Male", "Female"]),
                            first_name=row["first_name"],
                            last_name=row["last_name"],
                            middle_name=middle_name,
                            date_of_birth=row["date_of_birth"],
                            date_of_admission=row["date_of_admission"],
                            blood_type=row["blood_group"]
                        )
                    except IntegrityError:
                        studentID = generate_random_string(length=8) + get_initials(
                                row["first_name"], middle_name, row["last_name"]
                        ) + str(datetime.today().year)
                        student = Student.objects.create(
                            student_id=studentID,
                            gender=random.choice(["Male", "Female"]),
                            first_name=row["first_name"],
                            last_name=row["last_name"],
                            middle_name=middle_name,
                            date_of_birth=row["date_of_birth"],
                            date_of_admission=row["date_of_admission"],
                            blood_type=row["blood_group"]
                        )

                for item in fee_group:
                    fee_obj, _ = Fee.objects.get_or_create(
                            academic_year=acad_year,
                            academic_term=acad_term,
                            amount=item["amount"],
                            name=item["fees included"]
                        )
                    fee_group_obj, _ = StudentFeeGroup.objects.get_or_create(
                        name=item["fee group name"]
                    )
                    fee_group_obj.fees.add(fee_obj)
                student_fee_assigned = None
                if row["fee_group"] is not None or row["fee_group"] != "":
                    student_fee_assigned, _ = StudentFeeGroup.objects.get_or_create(
                        name__icontains=row["fee_group"]
                    )
                student_class, _ = StudentClass.objects.get_or_create(
                    student=student,
                    academic_year=acad_year,
                    student_class=my_class,
                    fee_assigned=student_fee_assigned
                )
        self.stdout.write(
                self.style.SUCCESS(
                    'Successfully created Students'
                    )
                )


def extract_excel_data(filename: str = "Updated HHA Enr.xlsx"):
    """Extract data from excel into json"""
    final_dict = {}
    fee_group_dict = {}
    dataframe = openpyxl.load_workbook(filename)

    name_list = dataframe.sheetnames
    for sheet_name in name_list:
        if sheet_name == "Fee Group":
            fee_sheet = dataframe[sheet_name]
            fee_headers = [cell.value.lower() for cell in fee_sheet[1]]
            fee_sheet_data = []
            for fee_row in fee_sheet.iter_rows(min_row=2, values_only=True):
                if all(ff is None for ff in fee_row):
                    pass
                else:
                    if not any(
                        sub in fr for fr in fee_row for sub in name_list if fr is not None
                        if type(fr) == str
                            ):
                        row_dict = dict(zip(fee_headers, fee_row))
                        fee_sheet_data.append(row_dict)
            fee_group_dict = fee_sheet_data
        else:
            sheet = dataframe[sheet_name]
            headers = [cell.value.lower() for cell in sheet[1]]
            sheet_data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if all(ff is None for ff in row):
                    pass
                else:
                    if not any(
                        sub in fr for fr in row for sub in name_list if fr is not None
                        if type(fr) == str
                            ):
                        row_dict = dict(zip(headers, row))
                        sheet_data.append(row_dict)
            final_dict[sheet_name] = sheet_data
    return final_dict, fee_group_dict
